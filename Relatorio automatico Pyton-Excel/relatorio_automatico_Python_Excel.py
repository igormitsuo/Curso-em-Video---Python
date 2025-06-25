
#Relatorio diario- Excel com Python

from openpyxl import Workbook 
from openpyxl import load_workbook
from time import sleep
from datetime import datetime


def menu():
 print('''
********** RELATORIO DIARIO DE ATIVIDADES **********

****** ESCOLHA A OPERAÇÃO DESEJADA ******
[0] SAIR
[1] CRIAR NOVA ABA
[2] PRENCHER RELATORIO
[3] EDITAR RELATORIO
[4] EXCLUIR RELATORIO
[5] EXCLUIR ABA
[6] VER ABAS CRIADAS
       
**************************************
''')
 
 valor=int(input("Entre com a opção \n"))
 if valor not in [0,1,2,3,4,5,6]:
    print("## OPÇÃO INVALIDA: DIGITE UM VALIDO##")
    sleep(1)
 return (valor)

def criar_nova_aba ():
 #arquivo=load_workbook ("relatorio_07-25.xlsx")#<<<<<<<AQUI, SALVAR O NOME DO ARQUIVO COPIADO
 global nome_aba
 global linha
 for aba in arquivo:
   print(aba)
 nome_aba=str(input("Digite o nome da nova aba\n"))
 nova_aba= arquivo.create_sheet(nome_aba)
 arquivo.active=nova_aba
 linha=arquivo[f'{nome_aba}']
 linha.append(["ITEM:","EQUIPAMENTO:","ÁREA:","DATA","HORA DE INICIO","HORA FIM","TOTAL","DESCRIÇÃO DA ATIVIDADE:","EXECUTANTES:","STATUS"])
 arquivo.save(f'{"relatorio_07-25.xlsx"}')#<<<<<<<AQUI, SALVAR O NOME DO ARQUIVO COPIADO

def preencher_relatorio(): 
 item=0
 while True:
  linha=arquivo[f'{nome_aba}']
  print(linha.max_row)
  item+=1
  equipamento=str(input("Digite o nome do equipamento\n"))
  área=str(input("Digite o nome da área\n"))
 # dt_inicio =input("Digite a data e hora de inicio da realização da atividade no formato Formato (DD\MM\AAAA HH:MM): \n")
 # dt_fim = input("Digite a data e hora de termino da realização da atividade no formato Formato (DD\MM\AAAA HH:MM): \n")
 # inicio = datetime.strptime(dt_inicio, "%d\%m\%Y %H:%M")
 # fim = datetime.strptime(dt_fim, "%d\%m\%Y %H:%M")
 # total = fim - inicio
  #print("Tempo total da atividade:", total)
  descrição=str(input("Digite a descrição da atividade\n"))
  executentes=str(input("Digite os nomes dos particitantes\n"))
  #status=input("digite o status OK OU NOK\n")
 # linha.max_row=linha.append([item,equipamento, área,dt_inicio,dt_fim,inicio,fim,total,descrição,executentes,status])
  linha.append([item,equipamento, área,descrição,executentes])
  arquivo.save(f'{"relatorio_07-25.xlsx"}')
  seleção=str(input('Deseja acrescentar nova atividade ? S \ N \n')).upper().strip()
  if seleção== "N":
   break
 arquivo.save(f'{"relatorio_07-25.xlsx"}')

def edição():
 print('''
********** EDIÇÃO DIARIO DE ATIVIDADES **********

****** ESCOLHA A OPERAÇÃO DESEJADA ******
[0] SAIR
[1] EDITAR DATA
[2] EDITAR CAMPO
[3] ADICIONAR LINHA
[4] EXCLUIR LINHA
       
**************************************
''')
 valor=int(input("Entre com a opção \n"))
 if valor not in [0,1,2,3,4,5,6]:
    print("## OPÇÃO INVALIDA: DIGITE UM VALIDO##")
    sleep(1)
 return (valor)



def editar_relatorio():
 arquivo=load_workbook ("relatorio_07-25.xlsx")
 global edt_aba
 
 edt_aba=(input("Informe qual data deseja editar\n"))
 edt_aba=arquivo.active
 while True:
  opção=edição()
  if opção==0:
   print("OPÇÃO SAIR")
   sleep(2)
   break 
  elif opção==1:
   nova_data=input("digite a nova data\n")
   edt_aba.title=nova_data
   arquivo.save(f'{"relatorio_07-25.xlsx"}')



while True:
 opção=menu()
 if opção==0:
  print("OPÇÃO SAIR")
  sleep(2)
  break
 
 elif opção==1:
  arquivo=load_workbook ("relatorio_07-25.xlsx")#<<<<<<<AQUI, SALVAR O NOME DO ARQUIVO COPIADO
  criar_nova_aba ()
 elif opção==2:
  preencher_relatorio()
 elif opção==3:
  editar_relatorio()
 